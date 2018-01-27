'''
Author: Paco Kwon (권해찬)
This program is written to automate the insertion of donation item data
This program needs an excel file, which matches the requirements stated below
'''
import mysql.connector
from mysql.connector import errorcode
from openpyxl import load_workbook


def create_database(cursor):
    try:
        cursor.execute("CREATE DATABASE {}\n".format(DB_NAME))
    except mysql.connector.Error as err:
        print("Failed creating database: {}".format(DB_NAME))
        exit(1)


config = {
    'user': 'kmlaonline',
    'password': 'n7h4eYWJ7vW59tT8'
}

DB_NAME = 'kmlaonline'

TABLES = {}
TABLES['donation_test'] = (
    "CREATE TABLE `donation_test` ("
    "  `n_num` int(11),"
    "  `n_category` int(11),"
    "  `s_title` text,"
    "  `n_who` bigint(20) NOT NULL,"
    "  `s_status` text,"
    "  `s_type` text,"
    "  `s_owner` text"
    ") ENGINE=InnoDB")

add_item = ("INSERT INTO donation_test "
            "(n_num, n_category, s_title, s_status, s_type, s_owner) "
            "VALUES (%s, %s, %s, %s, %s, %s)")

if __name__ == '__main__':
    # load excel file
    wb = load_workbook('donations.xlsx')

    cnx = mysql.connector.connect(**config)
    cursor = cnx.cursor()
    ''' connect to database '''
    try:
        cnx.database = DB_NAME
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_BAD_DB_ERROR:
            create_database(cursor)
            cnx.database = DB_NAME
        else:
            print(err)
            exit(1)
    ''' create table in TABLES '''
    for name, ddl in TABLES.items():
        try:
            print("Creating table {}".format(name), end='')
            cursor.execute(ddl)
        except mysql.connector.Error as err:
            if err.errno == errorcode.ER_TABLE_EXISTS_ERROR:
                print("\nalready exists.")
            else:
                print(err.msg)
        else:
            print("OK")
    '''
    n_num
    n_category
    ***excel file prerequisites***
    col 1: s_title
    col 2: s_status (status of daily)
    col 3: s_type (hanbok colour)
    col 4: s_owner (owner of hanbok)
    '''
    sheets = wb.get_sheet_names()
    for sheet in sheets:
        ws = wb[sheet]
        i = 1
        while ws.cell(row=i, column=1).value is not None:
            n_num = i
            n_category = sheets.index(sheet) + 1
            s_title = ws.cell(row=i, column=1).value
            s_status = ws.cell(row=i, column=2).value if \
                ws.cell(row=i, column=2).value is not None else ""
            s_type = ws.cell(row=i, column=3).value if \
                ws.cell(row=i, column=3).value is not None else ""
            s_owner = ws.cell(row=i, column=4).value if \
                ws.cell(row=i, column=4).value is not None else ""
            cursor.execute(add_item, (n_num, n_category, s_title, s_status, s_type, s_owner))
            i += 1

    cnx.commit()
    cursor.close()
    cnx.close()
