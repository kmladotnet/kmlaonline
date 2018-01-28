from openpyxl import load_workbook


def insertQuotes(msg):
    i = 0
    length = len(msg)
    while i < length:
        if msg[i] == "'":
            msg = msg[:i] + "'" + msg[i:]
            i += 1
        i += 1

    return msg


add_item = ("INSERT INTO donation_test "
            "(n_num, n_category, s_title, s_status, s_type, s_owner) "
            "VALUES ({}, {}, '{}', '{}', '{}', '{}')\n")

if __name__ == "__main__":
    f = open('donation_output.txt', 'w')
    wb = load_workbook('donations.xlsx')
    sheets = wb.get_sheet_names()
    for sheet in sheets:
        ws = wb[sheet]
        i = 1
        while ws.cell(row=i, column=1).value is not None:
            n_num = i
            n_category = sheets.index(sheet)
            s_title = insertQuotes(ws.cell(row=i, column=1).value) if \
                ws.cell(row=i, column=1).value is not None else ""
            s_status = insertQuotes(ws.cell(row=i, column=2).value) if \
                ws.cell(row=i, column=2).value is not None else ""
            s_type = insertQuotes(ws.cell(row=i, column=3).value) if \
                ws.cell(row=i, column=3).value is not None else ""
            s_owner = insertQuotes(ws.cell(row=i, column=4).value) if \
                ws.cell(row=i, column=4).value is not None else ""

            # print(add_item.format(n_num, n_category, s_title, s_status, s_type, s_owner))
            f.write(add_item.format(n_num, n_category, s_title, s_status, s_type, s_owner))
            i += 1
    f.close()
