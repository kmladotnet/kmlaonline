from openpyxl import load_workbook, Workbook


def chartoNum(char):
    if len(char) == 1:
        return ord(char.upper()) - 65 + 1
    elif len(char) == 2:
        char = char.upper()
        return chartoNum(char[0]) * 26 + chartoNum(char[1])


wb = load_workbook('donations.xlsx')
sheets = wb.get_sheet_names()

for sheet in sheets:
    ws = wb[sheet]
    print("SHEET: {}".format(sheet))
    i = 1
    print(ws.cell(row=i, column=1).value)
    while ws.cell(row=i, column=1).value is not None:
        print(ws.cell(row=i, column=1).value)
        i += 1
